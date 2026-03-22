#!/usr/bin/env python3
"""
Import Growatt PV Inverter Modbus RS485 RTU Protocol V3.14 into protocol database.

Legacy protocol for older grid-tied PV-only string inverters (pre-2017).
No battery or grid monitoring. Separate protocol family from Protocol II / VPP.

Source: Growatt-PV-Inverter-Modbus-RS485-RTU-Protocol-V3-14.xlsx
Sheet:  Table 1 (556 rows x 30 cols)

Verified column layout (1-indexed):
  Col  1: Register address (int) OR section marker text
  Col  3: Variable name
  Col  9: Description (may span continuation rows)
  Col 15: Customer Write / access type
  Col 17: Valid value range

Verified section boundaries (from direct inspection):
  Row 38: Header row ('Register NO.', 'Variable Name', etc.)
  Row 39: First holding register (address 0)
  Row 249: '4.2 Input Reg' section marker (contains 'input')
  Row 250: First input register
  Row 489: Last row with actual register address in col 1
             (rows 490+ are footnote tables with fault code lists)

Applicable models: Legacy Growatt PV string inverters (1-3 phase, pre-2017).
Section markers stored verbatim in notes — no model relationships inferred.
"""

import openpyxl
import sqlite3
import sys
from pathlib import Path
from datetime import date

DB_PATH = Path(__file__).parent.parent / 'docs' / 'protocol_database.db'
EXCEL_PATH = Path(__file__).parent.parent / 'Protocols' / 'Growatt-PV-Inverter-Modbus-RS485-RTU-Protocol-V3-14.xlsx'

PROTOCOL_NAME = 'PV_Modbus_v3.14'
PROTOCOL_FULL = 'Growatt PV Inverter Modbus RS485 RTU Protocol V3.14'
PROTOCOL_VERSION = '3.14'
APPLICABLE_MODELS = 'Legacy PV string inverters (pre-2017)'

# Verified from direct inspection
HOLDING_START = 39
HOLDING_END = 248       # row before '4.2 Input Reg' marker at row 249
INPUT_START = 250       # first data row after input section header
INPUT_END = 489         # last row with a real register address in col 1

# Column indices (1-based)
COL_ADDR = 1
COL_NAME = 3
COL_DESC = 9
COL_ACCESS = 15
COL_RANGE = 17


def clean_text(value):
    if value is None:
        return None
    text = str(value).strip()
    text = ' '.join(text.split())
    return text if text else None


def is_section_marker(value):
    if not isinstance(value, str):
        return False
    vl = value.lower().strip()
    keywords = ('group', 'use for', 'note', 'register table', 'second', 'third',
                'fourth', 'fifth', 'sixth', 'fault', 'reserved', 'appendix')
    return any(k in vl for k in keywords)


def create_or_get_protocol(db):
    cur = db.cursor()
    cur.execute("SELECT id FROM protocols WHERE name = ?", (PROTOCOL_NAME,))
    row = cur.fetchone()
    if row:
        print(f"Protocol '{PROTOCOL_NAME}' already exists (ID {row[0]}) — skipping creation")
        return row[0]
    cur.execute("""
        INSERT INTO protocols (name, full_name, version, source_file, applicable_models, created_date, updated_date)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """, (PROTOCOL_NAME, PROTOCOL_FULL, PROTOCOL_VERSION, EXCEL_PATH.name,
          APPLICABLE_MODELS, date.today().isoformat(), date.today().isoformat()))
    db.commit()
    pid = cur.lastrowid
    print(f"Created protocol '{PROTOCOL_NAME}' with ID {pid}")
    return pid


def parse_section(rows, db, protocol_id, start_row, end_row, register_type):
    """
    Parse registers from preloaded rows within [start_row, end_row] (1-indexed).
    Section markers stored verbatim in notes.
    """
    cursor = db.cursor()
    current_section = None
    current_reg = None
    added = 0

    def flush(reg):
        nonlocal added
        if not reg:
            return
        try:
            cursor.execute("""
                INSERT OR IGNORE INTO registers
                (protocol_id, register_type, address, name, description, access, valid_range, notes)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                protocol_id, register_type,
                reg['address'], reg['name'], reg['description'],
                reg.get('access'), reg.get('valid_range'), reg['section'],
            ))
            added += 1
        except Exception as e:
            print(f"    Error inserting {register_type} reg {reg['address']}: {e}")

    for sheet_row in range(start_row, end_row + 1):
        row = rows[sheet_row - 1]

        addr_val = row[COL_ADDR - 1]
        name_val = row[COL_NAME - 1] if len(row) >= COL_NAME else None
        desc_val = row[COL_DESC - 1] if len(row) >= COL_DESC else None
        access_val = row[COL_ACCESS - 1] if len(row) >= COL_ACCESS else None
        range_val = row[COL_RANGE - 1] if len(row) >= COL_RANGE else None

        # Section marker
        if isinstance(addr_val, str) and is_section_marker(addr_val):
            current_section = clean_text(addr_val)
            continue

        # Register row
        if isinstance(addr_val, (int, float)) and 0 <= int(addr_val) <= 70000:
            flush(current_reg)
            addr_int = int(addr_val)
            current_reg = {
                'address': addr_int,
                'name': clean_text(name_val) or f'reg_{addr_int}',
                'description': clean_text(desc_val) or '',
                'access': clean_text(access_val),
                'valid_range': clean_text(range_val),
                'section': current_section,
            }
        elif current_reg:
            # Continuation row
            extra = clean_text(desc_val)
            if extra:
                current_reg['description'] = (current_reg['description'] + ' ' + extra).strip()
            if range_val and not current_reg.get('valid_range'):
                current_reg['valid_range'] = clean_text(range_val)

    flush(current_reg)
    db.commit()
    print(f"    Added {added} {register_type} registers")
    return added


def main():
    print("=" * 80)
    print("Growatt PV Modbus V3.14 Import")
    print("=" * 80)

    if not EXCEL_PATH.exists():
        print(f"ERROR: {EXCEL_PATH}")
        return 1
    if not DB_PATH.exists():
        print(f"ERROR: {DB_PATH}")
        return 1

    print(f"\nExcel: {EXCEL_PATH.name}")
    print(f"DB:    {DB_PATH.name}")
    print(f"\nPreloading into memory...")

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True, read_only=True)
    ws = wb.worksheets[0]
    print(f"  Sheet: {ws.max_row} rows x {ws.max_column} cols")

    rows = list(ws.iter_rows(max_row=INPUT_END + 5, values_only=True))
    wb.close()
    print(f"  Loaded {len(rows)} rows into memory")

    db = sqlite3.connect(DB_PATH)
    protocol_id = create_or_get_protocol(db)

    print(f"\nImporting holding registers (rows {HOLDING_START}–{HOLDING_END})...")
    holding_count = parse_section(rows, db, protocol_id, HOLDING_START, HOLDING_END, 'holding')

    print(f"\nImporting input registers (rows {INPUT_START}–{INPUT_END})...")
    input_count = parse_section(rows, db, protocol_id, INPUT_START, INPUT_END, 'input')

    db.close()

    print("\n" + "=" * 80)
    print("Import Summary:")
    print(f"  Protocol:          {PROTOCOL_NAME} (ID: {protocol_id})")
    print(f"  Holding registers: {holding_count}")
    print(f"  Input registers:   {input_count}")
    print(f"  Total:             {holding_count + input_count}")
    print("=" * 80)
    return 0


if __name__ == '__main__':
    sys.exit(main())
