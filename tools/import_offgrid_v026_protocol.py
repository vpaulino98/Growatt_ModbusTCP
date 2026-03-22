#!/usr/bin/env python3
"""
Import Growatt OffGrid Modbus RS485/RS232 RTU Protocol V0.26 into protocol database.

Covers SPF off-grid and hybrid inverters only — a distinct protocol family with
separate register ranges from Protocol II.

Source: OffGrid-Modbus-RS485-RS232-RTU-Protocol-V0-26.xlsx
Sheet:  Table 1 (1014 rows x 70 cols)

Verified column layout (1-indexed):
  Col  1: Register address (int) OR section marker text
  Col  7: Variable name
  Col 21: Description (may span continuation rows)
  Col 33: Customer Write access
  Col 37: Valid value range

Verified section boundaries:
  Row 55: Header row
  Row 56: First holding register (address 0)
  Row 437: '4.2 Input Reg' section marker
  Row 438: First input register
  Row 1011: '5 Set address' — end of register data

Applicable models: SPF off-grid and hybrid inverters only.
Section markers stored verbatim in notes — no model relationships inferred.
"""

import openpyxl
import sqlite3
import sys
from pathlib import Path
from datetime import date

DB_PATH = Path(__file__).parent.parent / 'docs' / 'protocol_database.db'
EXCEL_PATH = Path(__file__).parent.parent / 'Protocols' / 'OffGrid-Modbus-RS485-RS232-RTU-Protocol-V0-26.xlsx'

PROTOCOL_NAME = 'OffGrid_RTU_v0.26'
PROTOCOL_FULL = 'Growatt OffGrid Modbus RS485/RS232 RTU Protocol V0.26'
PROTOCOL_VERSION = '0.26'
APPLICABLE_MODELS = 'SPF'

# Verified from direct inspection
HOLDING_START = 56
HOLDING_END = 436       # row before '4.2 Input Reg' marker
INPUT_START = 438       # first data row after input section header
INPUT_END = 1010        # last row with register integer in col 1

# Column indices (1-based)
COL_ADDR = 1
COL_NAME = 7
COL_DESC = 21
COL_ACCESS = 33
COL_RANGE = 37


def clean_text(value):
    if value is None:
        return None
    text = str(value).strip()
    text = ' '.join(text.split())
    return text if text else None


def is_section_marker(value):
    if not isinstance(value, str):
        return False
    vl = value.lower().strip()
    keywords = ('group', 'use for', 'note', 'register table', 'bms', 'battery',
                'ac output', 'ac input', 'solar', 'utility', 'second', 'third',
                'fourth', 'fifth', 'sixth', 'fault', 'warning', 'status',
                'set address', 'reserved')
    return any(k in vl for k in keywords)


def create_or_get_protocol(db):
    cur = db.cursor()
    cur.execute("SELECT id FROM protocols WHERE name = ?", (PROTOCOL_NAME,))
    row = cur.fetchone()
    if row:
        print(f"Protocol '{PROTOCOL_NAME}' already exists (ID {row[0]}) — skipping creation")
        return row[0]
    cur.execute("""
        INSERT INTO protocols (name, full_name, version, source_file, applicable_models, created_date, updated_date)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """, (PROTOCOL_NAME, PROTOCOL_FULL, PROTOCOL_VERSION, EXCEL_PATH.name,
          APPLICABLE_MODELS, date.today().isoformat(), date.today().isoformat()))
    db.commit()
    pid = cur.lastrowid
    print(f"Created protocol '{PROTOCOL_NAME}' with ID {pid}")
    return pid


def parse_section(rows, db, protocol_id, start_row, end_row, register_type):
    """
    Parse registers from preloaded rows. Section markers stored verbatim in notes.
    rows is 0-indexed; sheet rows are 1-indexed.
    """
    cursor = db.cursor()
    current_section = None
    current_reg = None
    added = 0

    def flush(reg):
        nonlocal added
        if not reg:
            return
        try:
            cursor.execute("""
                INSERT OR IGNORE INTO registers
                (protocol_id, register_type, address, name, description, access, valid_range, notes)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                protocol_id, register_type,
                reg['address'], reg['name'], reg['description'],
                reg.get('access'), reg.get('valid_range'), reg['section'],
            ))
            added += 1
        except Exception as e:
            print(f"    Error inserting {register_type} reg {reg['address']}: {e}")

    for sheet_row in range(start_row, end_row + 1):
        row = rows[sheet_row - 1]

        addr_val = row[COL_ADDR - 1]
        name_val = row[COL_NAME - 1] if len(row) >= COL_NAME else None
        desc_val = row[COL_DESC - 1] if len(row) >= COL_DESC else None
        access_val = row[COL_ACCESS - 1] if len(row) >= COL_ACCESS else None
        range_val = row[COL_RANGE - 1] if len(row) >= COL_RANGE else None

        # Section/group marker
        if isinstance(addr_val, str) and is_section_marker(addr_val):
            current_section = clean_text(addr_val)
            continue

        # Actual register row
        if isinstance(addr_val, (int, float)) and 0 <= int(addr_val) <= 70000:
            flush(current_reg)
            addr_int = int(addr_val)
            current_reg = {
                'address': addr_int,
                'name': clean_text(name_val) or f'reg_{addr_int}',
                'description': clean_text(desc_val) or '',
                'access': clean_text(access_val),
                'valid_range': clean_text(range_val),
                'section': current_section,
            }
        elif current_reg:
            # Continuation row — collect description/range fragments
            extra = clean_text(desc_val) or clean_text(name_val)
            if extra:
                current_reg['description'] = (current_reg['description'] + ' ' + extra).strip()
            if range_val and not current_reg.get('valid_range'):
                current_reg['valid_range'] = clean_text(range_val)

    flush(current_reg)
    db.commit()
    print(f"    Added {added} {register_type} registers")
    return added


def main():
    print("=" * 80)
    print("Growatt OffGrid RTU Protocol V0.26 Import")
    print("=" * 80)

    if not EXCEL_PATH.exists():
        print(f"ERROR: {EXCEL_PATH}")
        return 1
    if not DB_PATH.exists():
        print(f"ERROR: {DB_PATH}")
        return 1

    print(f"\nExcel: {EXCEL_PATH.name}")
    print(f"DB:    {DB_PATH.name}")
    print(f"\nPreloading into memory...")

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True, read_only=True)
    ws = wb.worksheets[0]
    print(f"  Sheet: {ws.max_row} rows x {ws.max_column} cols")

    rows = list(ws.iter_rows(max_row=INPUT_END + 5, values_only=True))
    wb.close()
    print(f"  Loaded {len(rows)} rows into memory")

    db = sqlite3.connect(DB_PATH)
    protocol_id = create_or_get_protocol(db)

    print(f"\nImporting holding registers (rows {HOLDING_START}–{HOLDING_END})...")
    holding_count = parse_section(rows, db, protocol_id, HOLDING_START, HOLDING_END, 'holding')

    print(f"\nImporting input registers (rows {INPUT_START}–{INPUT_END})...")
    input_count = parse_section(rows, db, protocol_id, INPUT_START, INPUT_END, 'input')

    db.close()

    print("\n" + "=" * 80)
    print("Import Summary:")
    print(f"  Protocol:          {PROTOCOL_NAME} (ID: {protocol_id})")
    print(f"  Holding registers: {holding_count}")
    print(f"  Input registers:   {input_count}")
    print(f"  Total:             {holding_count + input_count}")
    print("=" * 80)
    return 0


if __name__ == '__main__':
    sys.exit(main())
