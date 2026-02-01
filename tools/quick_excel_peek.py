#!/usr/bin/env python3
"""Quick peek at Excel file - export first N rows to show structure"""

import sys
from pathlib import Path
from openpyxl import load_workbook

def export_rows(filepath, num_rows=100):
    """Export first N rows from Excel to show structure"""
    print(f"\nFile: {filepath.name}")
    print("=" * 120)

    wb = load_workbook(filepath, data_only=True, read_only=False)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\nSheet: {sheet_name} ({ws.max_row} rows × {ws.max_column} columns)")
        print("-" * 120)

        # Print first num_rows
        for row_idx in range(1, min(num_rows + 1, ws.max_row + 1)):
            row_vals = []
            for col_idx in range(1, min(15, ws.max_column + 1)):  # First 15 columns
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is not None:
                    val_str = str(val).replace('\n', ' ')[:60]  # Truncate long values
                    row_vals.append(f"{col_idx}:{val_str}")

            if row_vals:
                print(f"Row {row_idx:3d}: {' | '.join(row_vals)}")

    wb.close()

if __name__ == '__main__':
    protocols_dir = Path('/home/user/Growatt_ModbusTCP/Protocols')

    # VPP file
    print("\n" + "="*120)
    print("VPP 2.03 PROTOCOL - First 100 rows")
    print("="*120)
    vpp_file = protocols_dir / 'GI-BK-E060_GROWATT.VPP.COMMUNICATION.PROTOCOL.OF.INVERTER_V2.03.xlsx'
    if vpp_file.exists():
        export_rows(vpp_file, 100)
