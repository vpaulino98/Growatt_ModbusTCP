#!/usr/bin/env python3
"""
Import Growatt Modbus RTU Protocol II V1.24 into protocol database.

This script handles the single-sheet format with misaligned columns and
multi-row descriptions. It extracts holding registers and input registers,
tracking which inverter types each register applies to.
"""

import openpyxl
import sqlite3
import sys
from pathlib import Path
from datetime import date

# Database path
DB_PATH = Path(__file__).parent.parent / 'docs' / 'protocol_database.db'
EXCEL_PATH = Path(__file__).parent.parent / 'Protocols' / 'Growatt-Inverter-Modbus-RTU-Protocol-II-V1-24-English-new.xlsx'

# Column mappings for holding registers
HOLDING_REG_COLS = {
    'address': 1,
    'name': 4,
    'description': 11,
    'access': 20,
    'value_range': 23,
    'unit': 28,
}

# Column mappings for input registers
INPUT_REG_COLS = {
    'address': 1,
    'name': 4,  # Actually appears in col 4 despite header saying col 3
    'description': 16,
}

# Section boundaries (approximate)
HOLDING_START_ROW = 41
HOLDING_END_ROW = 704
INPUT_START_ROW = 708


def create_protocol(db, name, full_name, version, source_file):
    """Create or get protocol entry."""
    cursor = db.cursor()

    # Check if protocol exists
    cursor.execute("SELECT id FROM protocols WHERE name = ?", (name,))
    result = cursor.fetchone()

    if result:
        print(f"Protocol '{name}' already exists with ID {result[0]}")
        return result[0]

    # Create new protocol
    cursor.execute("""
        INSERT INTO protocols (name, full_name, version, source_file, created_date, updated_date)
        VALUES (?, ?, ?, ?, ?, ?)
    """, (name, full_name, version, source_file, date.today(), date.today()))

    protocol_id = cursor.lastrowid
    print(f"Created protocol '{name}' with ID {protocol_id}")
    return protocol_id


def parse_holding_registers(sheet, db, protocol_id):
    """Parse holding registers from sheet, handling multi-row entries."""
    cursor = db.cursor()

    current_reg = None
    current_section = None  # Track which models this section applies to
    registers_added = 0

    for row_idx in range(HOLDING_START_ROW, HOLDING_END_ROW + 1):
        # Get values from mapped columns
        address_val = sheet.cell(row=row_idx, column=HOLDING_REG_COLS['address']).value
        name_val = sheet.cell(row=row_idx, column=HOLDING_REG_COLS['name']).value
        desc_val = sheet.cell(row=row_idx, column=HOLDING_REG_COLS['description']).value
        access_val = sheet.cell(row=row_idx, column=HOLDING_REG_COLS['access']).value
        range_val = sheet.cell(row=row_idx, column=HOLDING_REG_COLS['value_range']).value
        unit_val = sheet.cell(row=row_idx, column=HOLDING_REG_COLS['unit']).value

        # Check for section markers (e.g., "Use for TL-X and TL-XH")
        if isinstance(address_val, str) and ('use for' in address_val.lower() or 'group' in address_val.lower()):
            current_section = address_val
            print(f"  Section marker at row {row_idx}: {current_section[:60]}")
            continue

        # If we have an address (register number), start a new register entry
        if isinstance(address_val, int):
            # Save previous register if exists
            if current_reg:
                try:
                    cursor.execute("""
                        INSERT OR IGNORE INTO registers
                        (protocol_id, register_type, address, name, description, access, valid_range, unit, notes)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (
                        protocol_id,
                        'holding',
                        current_reg['address'],
                        current_reg['name'],
                        current_reg['description'],
                        current_reg['access'],
                        current_reg['range'],
                        current_reg['unit'],
                        current_reg['section']
                    ))
                    registers_added += 1
                except Exception as e:
                    print(f"    Error inserting register {current_reg['address']}: {e}")

            # Start new register
            current_reg = {
                'address': address_val,
                'name': clean_text(name_val) if name_val else f'reg_{address_val}',
                'description': clean_text(desc_val) if desc_val else '',
                'access': clean_text(access_val) if access_val else 'R',
                'range': clean_text(range_val) if range_val else None,
                'unit': clean_text(unit_val) if unit_val else None,
                'section': current_section
            }

        # If no address but we have a description, append to current register's description
        elif current_reg and desc_val and not isinstance(address_val, int):
            desc_text = clean_text(desc_val)
            if desc_text:
                current_reg['description'] += ' ' + desc_text

    # Save last register
    if current_reg:
        try:
            cursor.execute("""
                INSERT OR IGNORE INTO registers
                (protocol_id, register_type, address, name, description, access, valid_range, unit, notes)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                protocol_id,
                'holding',
                current_reg['address'],
                current_reg['name'],
                current_reg['description'],
                current_reg['access'],
                current_reg['range'],
                current_reg['unit'],
                current_reg['section']
            ))
            registers_added += 1
        except Exception as e:
            print(f"    Error inserting last register: {e}")

    db.commit()
    print(f"  Added {registers_added} holding registers")
    return registers_added


def parse_input_registers(sheet, db, protocol_id):
    """Parse input registers from sheet, handling multi-row entries."""
    cursor = db.cursor()

    current_reg = None
    current_section = None
    registers_added = 0

    # Find where input registers actually end
    last_row = sheet.max_row

    for row_idx in range(INPUT_START_ROW, last_row + 1):
        # Get values from mapped columns
        address_val = sheet.cell(row=row_idx, column=INPUT_REG_COLS['address']).value
        name_val = sheet.cell(row=row_idx, column=INPUT_REG_COLS['name']).value
        desc_val = sheet.cell(row=row_idx, column=INPUT_REG_COLS['description']).value

        # Check for section markers
        if isinstance(address_val, str) and ('use for' in address_val.lower() or 'group' in address_val.lower()):
            current_section = address_val
            print(f"  Section marker at row {row_idx}: {current_section[:60]}")
            continue

        # Stop if we hit the end notice
        if isinstance(address_val, str) and 'app user' in address_val.lower():
            print(f"  Reached end of input registers at row {row_idx}")
            break

        # If we have an address (register number), start a new register entry
        if isinstance(address_val, int):
            # Save previous register if exists
            if current_reg:
                try:
                    cursor.execute("""
                        INSERT OR IGNORE INTO registers
                        (protocol_id, register_type, address, name, description, notes)
                        VALUES (?, ?, ?, ?, ?, ?)
                    """, (
                        protocol_id,
                        'input',
                        current_reg['address'],
                        current_reg['name'],
                        current_reg['description'],
                        current_reg['section']
                    ))
                    registers_added += 1
                except Exception as e:
                    print(f"    Error inserting register {current_reg['address']}: {e}")

            # Start new register
            current_reg = {
                'address': address_val,
                'name': clean_text(name_val) if name_val else f'input_reg_{address_val}',
                'description': clean_text(desc_val) if desc_val else '',
                'section': current_section
            }

        # If no address but we have a description, append to current register's description
        elif current_reg and desc_val and not isinstance(address_val, int):
            desc_text = clean_text(desc_val)
            if desc_text:
                current_reg['description'] += ' ' + desc_text

    # Save last register
    if current_reg:
        try:
            cursor.execute("""
                INSERT OR IGNORE INTO registers
                (protocol_id, register_type, address, name, description, notes)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (
                protocol_id,
                'input',
                current_reg['address'],
                current_reg['name'],
                current_reg['description'],
                current_reg['section']
            ))
            registers_added += 1
        except Exception as e:
            print(f"    Error inserting last register: {e}")

    db.commit()
    print(f"  Added {registers_added} input registers")
    return registers_added


def clean_text(text):
    """Clean text by removing extra whitespace and newlines."""
    if not text:
        return None

    if isinstance(text, str):
        # Replace newlines with spaces, collapse multiple spaces
        cleaned = ' '.join(text.split())
        return cleaned if cleaned else None

    return str(text)


def main():
    print("=" * 80)
    print("Growatt Modbus RTU Protocol II V1.24 Import")
    print("=" * 80)

    # Check if files exist
    if not EXCEL_PATH.exists():
        print(f"Error: Excel file not found: {EXCEL_PATH}")
        return 1

    if not DB_PATH.exists():
        print(f"Error: Database not found: {DB_PATH}")
        print("Please run import_vpp_protocol.py first to create the database.")
        return 1

    print(f"\nExcel file: {EXCEL_PATH}")
    print(f"Database: {DB_PATH}")

    # Load workbook
    print("\nLoading Excel workbook...")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True, read_only=True)
    sheet = wb['Table 1']
    print(f"  Loaded sheet 'Table 1' ({sheet.max_row} rows)")

    # Connect to database
    print("\nConnecting to database...")
    db = sqlite3.connect(DB_PATH)

    # Create protocol entry
    print("\nCreating protocol entry...")
    protocol_id = create_protocol(
        db,
        name='RTU_Protocol_II_v1.24',
        full_name='Growatt Inverter Modbus RTU Protocol II',
        version='1.24',
        source_file=str(EXCEL_PATH.name)
    )

    # Import holding registers
    print("\nImporting holding registers...")
    print(f"  Rows {HOLDING_START_ROW} to {HOLDING_END_ROW}")
    holding_count = parse_holding_registers(sheet, db, protocol_id)

    # Import input registers
    print("\nImporting input registers...")
    print(f"  Starting from row {INPUT_START_ROW}")
    input_count = parse_input_registers(sheet, db, protocol_id)

    # Close connections
    wb.close()
    db.close()

    print("\n" + "=" * 80)
    print("Import Summary:")
    print(f"  Protocol: RTU Protocol II v1.24 (ID: {protocol_id})")
    print(f"  Holding registers: {holding_count}")
    print(f"  Input registers: {input_count}")
    print(f"  Total registers: {holding_count + input_count}")
    print("=" * 80)

    return 0


if __name__ == '__main__':
    sys.exit(main())
