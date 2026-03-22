#!/usr/bin/env python3
"""
Import Growatt WIT Modbus RTU Protocol II V1.39 into protocol database.

This is the definitive Protocol II document — V1.39 is a verified superset of
V1.13, V1.20, and V1.24. It replaces those older versions.

Source: Growatt_WIT-Modbus-RTU-Protocol-II-V1.39.xlsx
Sheet:  Table 1 (2171 rows x 57 cols)

Verified column layout (1-indexed):
  Col  1: Register address (int) OR section marker text
  Col  4: Variable name
  Col  8: Description (may span many continuation rows)
  Col 20: Read/Write access

Verified section boundaries:
  Row 35: Header row
  Row 36: 'First group' section marker
  Row 37: First holding register (address 0)
  Row 863: '2.2 Input register' section marker
  Row 864: First input register (address 0)
  Row 2005: '4 Notice' — appendix/footnotes begin (no more registers)

Applicable models (from document intro row 32):
  TL-X/TL-XH (MIN type): 0~124, 3000~3124, 3125~3249, 3250~3374
  TL3-X (MAX/MID/MAC type): 0~124, 3000~3124, 3125~3249
  SPA/SPH (hybrid): 0~124, 1000~1124, 2000~2124, 3000~3124, 3125~3249
  WIT TL3: 0~124, 875~999, 3000~3124, 3125~3249, 8000~8139

Section markers (model-specific notes) are stored verbatim in the 'notes' field.
No model-register relationships are inferred — only what is explicitly stated.
"""

import openpyxl
import sqlite3
import sys
from pathlib import Path
from datetime import date

DB_PATH = Path(__file__).parent.parent / 'docs' / 'protocol_database.db'
EXCEL_PATH = Path(__file__).parent.parent / 'Protocols' / 'Growatt_WIT-Modbus-RTU-Protocol-II-V1.39.xlsx'

PROTOCOL_NAME = 'RTU_Protocol_II_v1.39'
PROTOCOL_FULL = 'Growatt Inverter Modbus RTU Protocol II V1.39 (WIT)'
PROTOCOL_VERSION = '1.39'
APPLICABLE_MODELS = 'WIT,WIS,SPH,SPA,MIN,MOD,MID,MAX,MAC'

# Verified from direct inspection
HEADER_ROW = 35
HOLDING_START = 37
HOLDING_END = 862       # row before '2.2 Input register' marker
INPUT_START = 864       # first data row after input section header
INPUT_END = 2004        # row before '4 Notice' appendix

# Column indices (1-based)
COL_ADDR = 1
COL_NAME = 4
COL_DESC = 8
COL_ACCESS = 20


def clean_text(value):
    if value is None:
        return None
    text = str(value).strip()
    text = ' '.join(text.split())
    return text if text else None


def is_section_marker(value):
    """True if this col-1 value is a section/group header rather than a register address."""
    if not isinstance(value, str):
        return False
    vl = value.lower().strip()
    keywords = ('group', 'use for', 'for tl', 'for wit', 'for sph', 'for spa',
                'note', 'register table', 'appendix', 'bms', 'apx', 'reserved',
                'second', 'third', 'fourth', 'fifth', 'sixth', 'seventh', 'eighth',
                'ninth', 'tenth', 'eleventh', 'twelfth', 'thirteenth')
    return any(k in vl for k in keywords)


def create_or_get_protocol(db):
    cur = db.cursor()
    cur.execute("SELECT id FROM protocols WHERE name = ?", (PROTOCOL_NAME,))
    row = cur.fetchone()
    if row:
        print(f"Protocol '{PROTOCOL_NAME}' already exists (ID {row[0]}) — skipping creation")
        return row[0]
    cur.execute("""
        INSERT INTO protocols (name, full_name, version, source_file, applicable_models, created_date, updated_date)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """, (PROTOCOL_NAME, PROTOCOL_FULL, PROTOCOL_VERSION, EXCEL_PATH.name,
          APPLICABLE_MODELS, date.today().isoformat(), date.today().isoformat()))
    db.commit()
    pid = cur.lastrowid
    print(f"Created protocol '{PROTOCOL_NAME}' with ID {pid}")
    return pid


def parse_section(rows, db, protocol_id, start_row, end_row, register_type):
    """
    Parse register rows from preloaded row data (0-indexed list).
    rows[i] corresponds to spreadsheet row (i+1).

    Section header text (explicitly written in col 1 of the source) is stored
    verbatim in the notes field — this is the only model-applicability link.
    """
    cursor = db.cursor()
    current_section = None
    current_reg = None
    added = 0

    def flush(reg):
        nonlocal added
        if not reg:
            return
        try:
            cursor.execute("""
                INSERT OR IGNORE INTO registers
                (protocol_id, register_type, address, name, description, access, notes)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (
                protocol_id, register_type,
                reg['address'], reg['name'], reg['description'],
                reg.get('access'), reg['section'],
            ))
            added += 1
        except Exception as e:
            print(f"    Error inserting {register_type} reg {reg['address']}: {e}")

    for sheet_row in range(start_row, end_row + 1):
        row = rows[sheet_row - 1]  # convert to 0-based index

        addr_val = row[COL_ADDR - 1]
        name_val = row[COL_NAME - 1]
        desc_val = row[COL_DESC - 1]
        access_val = row[COL_ACCESS - 1] if len(row) >= COL_ACCESS else None

        # Section/group marker (text in address column)
        if isinstance(addr_val, str) and is_section_marker(addr_val):
            current_section = clean_text(addr_val)
            continue

        # Actual register row
        if isinstance(addr_val, (int, float)) and 0 <= int(addr_val) <= 70000:
            flush(current_reg)
            addr_int = int(addr_val)
            current_reg = {
                'address': addr_int,
                'name': clean_text(name_val) or f'reg_{addr_int}',
                'description': clean_text(desc_val) or '',
                'access': clean_text(access_val),
                'section': current_section,
            }
        elif current_reg:
            # Continuation row — append description fragments
            for col_val in [desc_val, name_val]:
                if col_val is not None:
                    extra = clean_text(col_val)
                    if extra:
                        current_reg['description'] = (current_reg['description'] + ' ' + extra).strip()
                    break

    flush(current_reg)
    db.commit()
    print(f"    Added {added} {register_type} registers")
    return added


def main():
    print("=" * 80)
    print("Growatt RTU Protocol II V1.39 (WIT) Import")
    print("=" * 80)

    if not EXCEL_PATH.exists():
        print(f"ERROR: {EXCEL_PATH}")
        return 1
    if not DB_PATH.exists():
        print(f"ERROR: {DB_PATH}")
        return 1

    print(f"\nExcel: {EXCEL_PATH.name}")
    print(f"DB:    {DB_PATH.name}")
    print(f"\nPreloading {EXCEL_PATH.name} into memory...")

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True, read_only=True)
    ws = wb.worksheets[0]
    print(f"  Sheet: {ws.max_row} rows x {ws.max_column} cols")

    # Load all rows at once — fast sequential scan, avoids O(n²) cell lookups
    rows = list(ws.iter_rows(max_row=INPUT_END + 10, values_only=True))
    wb.close()
    print(f"  Loaded {len(rows)} rows into memory")

    db = sqlite3.connect(DB_PATH)
    protocol_id = create_or_get_protocol(db)

    print(f"\nImporting holding registers (rows {HOLDING_START}–{HOLDING_END})...")
    holding_count = parse_section(rows, db, protocol_id, HOLDING_START, HOLDING_END, 'holding')

    print(f"\nImporting input registers (rows {INPUT_START}–{INPUT_END})...")
    input_count = parse_section(rows, db, protocol_id, INPUT_START, INPUT_END, 'input')

    db.close()

    print("\n" + "=" * 80)
    print("Import Summary:")
    print(f"  Protocol:          {PROTOCOL_NAME} (ID: {protocol_id})")
    print(f"  Holding registers: {holding_count}")
    print(f"  Input registers:   {input_count}")
    print(f"  Total:             {holding_count + input_count}")
    print("=" * 80)
    return 0


if __name__ == '__main__':
    sys.exit(main())
