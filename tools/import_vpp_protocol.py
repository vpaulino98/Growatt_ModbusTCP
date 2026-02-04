#!/usr/bin/env python3
"""
Import VPP 2.03 Protocol from Excel to SQLite Database

This is a proof-of-concept import script demonstrating how clean the
restructured Excel file imports into a database.
"""

import sqlite3
from pathlib import Path
from openpyxl import load_workbook
from datetime import datetime


def create_database_schema(db):
    """Create database tables"""

    db.execute("""
        CREATE TABLE IF NOT EXISTS protocols (
            id INTEGER PRIMARY KEY,
            name TEXT NOT NULL UNIQUE,
            full_name TEXT,
            version TEXT,
            source_file TEXT,
            notes TEXT,
            created_date DATE,
            updated_date DATE
        )
    """)

    db.execute("""
        CREATE TABLE IF NOT EXISTS registers (
            id INTEGER PRIMARY KEY,
            protocol_id INTEGER,
            register_type TEXT,
            address INTEGER NOT NULL,
            name TEXT NOT NULL,
            parameter_name TEXT,
            description TEXT,
            data_type TEXT,
            unit TEXT,
            scale REAL,
            access TEXT,
            register_count INTEGER,
            valid_range TEXT,
            paired_with INTEGER,
            signed BOOLEAN,
            notes TEXT,

            FOREIGN KEY (protocol_id) REFERENCES protocols(id),
            UNIQUE(protocol_id, register_type, address)
        )
    """)

    db.execute("CREATE INDEX IF NOT EXISTS idx_registers_protocol ON registers(protocol_id)")
    db.execute("CREATE INDEX IF NOT EXISTS idx_registers_address ON registers(address)")
    db.execute("CREATE INDEX IF NOT EXISTS idx_registers_name ON registers(name)")

    db.execute("""
        CREATE TABLE IF NOT EXISTS dtc_codes (
            id INTEGER PRIMARY KEY,
            protocol_id INTEGER,
            dtc_code INTEGER NOT NULL,
            inverter_type TEXT,
            full_name TEXT,
            model_series TEXT,
            profile_name TEXT,
            notes TEXT,

            FOREIGN KEY (protocol_id) REFERENCES protocols(id),
            UNIQUE(dtc_code, protocol_id)
        )
    """)

    db.commit()


def create_protocol(db, name, full_name, version, source_file):
    """Insert protocol metadata"""

    cursor = db.execute("""
        INSERT OR REPLACE INTO protocols (name, full_name, version, source_file, created_date, updated_date)
        VALUES (?, ?, ?, ?, ?, ?)
    """, (name, full_name, version, source_file, datetime.now(), datetime.now()))

    db.commit()
    return cursor.lastrowid


def clean_header(text):
    """Clean header text (remove line breaks)"""
    if text:
        return str(text).replace('\n', ' ').replace('\r', ' ').strip()
    return None


def import_holding_registers(sheet, db, protocol_id):
    """Import holding registers from HOLDING REGISTERS sheet"""

    print(f"Importing holding registers...")

    # Row 2 has headers
    headers = {}
    for col_idx in range(1, sheet.max_column + 1):
        val = sheet.cell(row=2, column=col_idx).value
        if val:
            headers[col_idx] = clean_header(val)

    # Find key columns (by header text)
    col_no = next((k for k, v in headers.items() if 'N O' in v), None)
    col_name = next((k for k, v in headers.items() if 'Parameter name' in v), None)
    col_access = next((k for k, v in headers.items() if 'rite' in v), None)  # "Read/write"
    col_type = next((k for k, v in headers.items() if v == 'Type'), None)
    col_unit = next((k for k, v in headers.items() if v == 'Unit'), None)
    col_addr = next((k for k, v in headers.items() if v == 'Address'), None)
    col_count = next((k for k, v in headers.items() if 'umber' in v), None)  # "Number"
    col_range = next((k for k, v in headers.items() if v == 'Range'), None)

    print(f"  Column mapping: NO={col_no}, Name={col_name}, Access={col_access}, Type={col_type}, "
          f"Unit={col_unit}, Address={col_addr}, Count={col_count}, Range={col_range}")

    # Import data rows (starting from row 4)
    count = 0
    for row_idx in range(4, sheet.max_row + 1):
        # Get values
        no = sheet.cell(row=row_idx, column=col_no).value if col_no else None
        param_name = sheet.cell(row=row_idx, column=col_name).value if col_name else None
        access = sheet.cell(row=row_idx, column=col_access).value if col_access else None
        data_type = sheet.cell(row=row_idx, column=col_type).value if col_type else None
        unit = sheet.cell(row=row_idx, column=col_unit).value if col_unit else None
        address = sheet.cell(row=row_idx, column=col_addr).value if col_addr else None
        reg_count = sheet.cell(row=row_idx, column=col_count).value if col_count else None
        valid_range = sheet.cell(row=row_idx, column=col_range).value if col_range else None

        # Skip if no address (section headers or empty rows)
        if not address or not isinstance(address, (int, float)):
            continue

        # Clean values
        param_name = clean_header(param_name) if param_name else None
        access = clean_header(access) if access else None

        # Derive name from parameter_name (simplified - would need more logic)
        name = param_name.lower().replace(' ', '_').replace('(', '').replace(')', '') if param_name else None

        # Insert register
        db.execute("""
            INSERT OR REPLACE INTO registers
            (protocol_id, register_type, address, name, parameter_name, data_type,
             unit, access, register_count, valid_range)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (protocol_id, 'holding', int(address), name, param_name, data_type,
              unit, access, int(reg_count) if reg_count else 1, valid_range))

        count += 1

    db.commit()
    print(f"  Imported {count} holding registers")


def import_input_registers(sheet, db, protocol_id):
    """Import input registers from input registers sheet"""

    print(f"Importing input registers...")

    # Row 2 has headers
    headers = {}
    for col_idx in range(1, min(10, sheet.max_column + 1)):
        val = sheet.cell(row=2, column=col_idx).value
        if val:
            headers[col_idx] = clean_header(val)

    # Simple column mapping (1-8 are the standard columns)
    col_no = 1
    col_name = 2
    col_access = 3
    col_type = 4
    col_unit = 5
    col_addr = 6
    col_count = 7
    col_range = 8

    # Import data rows (starting from row 4)
    count = 0
    for row_idx in range(4, sheet.max_row + 1):
        # Get values
        address = sheet.cell(row=row_idx, column=col_addr).value

        # Skip if no address
        if not address or not isinstance(address, (int, float)):
            continue

        param_name = sheet.cell(row=row_idx, column=col_name).value
        access = sheet.cell(row=row_idx, column=col_access).value
        data_type = sheet.cell(row=row_idx, column=col_type).value
        unit = sheet.cell(row=row_idx, column=col_unit).value
        reg_count = sheet.cell(row=row_idx, column=col_count).value
        valid_range = sheet.cell(row=row_idx, column=col_range).value

        # Clean values
        param_name = clean_header(param_name) if param_name else None
        access = clean_header(access) if access else None

        # Derive name
        name = param_name.lower().replace(' ', '_').replace('(', '').replace(')', '') if param_name else None

        # Insert register
        db.execute("""
            INSERT OR REPLACE INTO registers
            (protocol_id, register_type, address, name, parameter_name, data_type,
             unit, access, register_count, valid_range)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (protocol_id, 'input', int(address), name, param_name, data_type,
              unit, access, int(reg_count) if reg_count else 1, valid_range))

        count += 1

    db.commit()
    print(f"  Imported {count} input registers")


def import_dtc_codes(sheet, db, protocol_id):
    """Import DTC codes from TABLES sheet"""

    print(f"Importing DTC codes...")

    # DTC table starts at row 1
    # Headers at row 2: Type (col 2), Full name (col 19), DTC code (col 44)
    count = 0
    for row_idx in range(3, 30):  # DTC table is roughly rows 3-30
        inv_type = sheet.cell(row=row_idx, column=2).value
        full_name = sheet.cell(row=row_idx, column=19).value
        dtc_code = sheet.cell(row=row_idx, column=44).value

        if dtc_code and isinstance(dtc_code, (int, float)):
            # Clean text
            inv_type = clean_header(inv_type) if inv_type else None
            full_name = clean_header(full_name) if full_name else None

            # Extract model series from type (e.g., "SPH/SPA" → "SPH")
            model_series = inv_type.split('/')[0] if inv_type and '/' in inv_type else inv_type

            db.execute("""
                INSERT OR REPLACE INTO dtc_codes
                (protocol_id, dtc_code, inverter_type, full_name, model_series)
                VALUES (?, ?, ?, ?, ?)
            """, (protocol_id, int(dtc_code), inv_type, full_name, model_series))

            count += 1

    db.commit()
    print(f"  Imported {count} DTC codes")


def main():
    """Main import process"""

    print("VPP 2.03 Protocol Import")
    print("=" * 80)

    # Paths
    excel_file = Path('/home/user/Growatt_ModbusTCP/Protocols/GI-BK-E060_GROWATT.VPP.COMMUNICATION.PROTOCOL.OF.INVERTER_V2.03.xlsx')
    db_file = Path('/home/user/Growatt_ModbusTCP/docs/protocol_database.db')

    # Load Excel file
    print(f"\nLoading {excel_file.name}...")
    wb = load_workbook(excel_file, data_only=True)
    print(f"  Sheets: {', '.join(wb.sheetnames)}")

    # Connect to database
    print(f"\nConnecting to database: {db_file.name}")
    db = sqlite3.connect(db_file)

    # Create schema
    print("Creating database schema...")
    create_database_schema(db)

    # Create protocol entry
    print("\nCreating protocol entry...")
    protocol_id = create_protocol(
        db,
        name='VPP_2.03',
        full_name='VPP Communication Protocol V2.03',
        version='2.03',
        source_file=excel_file.name
    )
    print(f"  Protocol ID: {protocol_id}")

    # Import data
    print("\nImporting data...")
    import_holding_registers(wb['HOLDING REGISTERS'], db, protocol_id)
    import_input_registers(wb['input registers'], db, protocol_id)
    import_dtc_codes(wb['TABLES'], db, protocol_id)

    # Summary
    print("\n" + "=" * 80)
    print("Import Summary:")
    cursor = db.execute("SELECT COUNT(*) FROM registers WHERE protocol_id=? AND register_type='holding'", (protocol_id,))
    print(f"  Holding registers: {cursor.fetchone()[0]}")
    cursor = db.execute("SELECT COUNT(*) FROM registers WHERE protocol_id=? AND register_type='input'", (protocol_id,))
    print(f"  Input registers: {cursor.fetchone()[0]}")
    cursor = db.execute("SELECT COUNT(*) FROM dtc_codes WHERE protocol_id=?", (protocol_id,))
    print(f"  DTC codes: {cursor.fetchone()[0]}")

    print(f"\n✓ Import complete! Database: {db_file}")

    db.close()
    wb.close()


if __name__ == '__main__':
    main()
