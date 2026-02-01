#!/usr/bin/env python3
"""
Analyze Growatt Protocol Excel files to understand structure
Scans entire document to find register tables, DTC tables, error codes, etc.
"""

import sys
from pathlib import Path
from openpyxl import load_workbook
from collections import defaultdict

def find_table_sections(ws):
    """Scan worksheet to find different table sections"""
    max_row = ws.max_row
    max_col = ws.max_column

    sections = []
    current_section = None

    print(f"\n    Scanning {max_row} rows for table patterns...")

    for row_idx in range(1, max_row + 1):
        # Get all non-empty cells in this row
        row_vals = []
        for col_idx in range(1, min(max_col + 1, 20)):  # Check first 20 columns
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                row_vals.append(str(val).strip())

        if not row_vals:
            continue

        row_text = ' '.join(row_vals).lower()

        # Look for section headers/markers
        section_keywords = {
            'register': ['register', 'addr', 'address', 'function code'],
            'dtc': ['dtc', 'device type code', 'device code'],
            'error': ['error code', 'fault code', 'alarm'],
            'input': ['input register', 'holding register', 'coil'],
            'table': ['table', 'list'],
        }

        # Check if this row looks like a table header
        is_header = False
        section_type = None

        for stype, keywords in section_keywords.items():
            if any(kw in row_text for kw in keywords):
                # Check if it has multiple column-like values
                if len(row_vals) >= 3:
                    is_header = True
                    section_type = stype
                    break

        if is_header:
            # Save previous section if exists
            if current_section:
                sections.append(current_section)

            current_section = {
                'type': section_type,
                'start_row': row_idx,
                'header_row': row_idx,
                'headers': row_vals,
                'sample_rows': []
            }
            print(f"      Found {section_type.upper()} table at row {row_idx}: {row_vals[:8]}")

        elif current_section and len(current_section['sample_rows']) < 5:
            # Collect sample data rows
            if len(row_vals) >= 3:  # Has enough columns to be data
                current_section['sample_rows'].append({
                    'row': row_idx,
                    'data': row_vals[:10]
                })

    # Add last section
    if current_section:
        sections.append(current_section)

    return sections

def analyze_excel_file(filepath):
    """Analyze structure of an Excel file"""
    print(f"\n{'='*80}")
    print(f"File: {filepath.name}")
    print(f"{'='*80}")

    try:
        wb = load_workbook(filepath, data_only=True, read_only=True)
        print(f"\nTotal sheets: {len(wb.sheetnames)}")

        for sheet_name in wb.sheetnames:
            print(f"\n  Sheet: '{sheet_name}'")
            ws = wb[sheet_name]

            max_row = ws.max_row
            max_col = ws.max_column
            print(f"    Dimensions: {max_row} rows × {max_col} columns")

            # Find table sections
            sections = find_table_sections(ws)

            print(f"\n    Found {len(sections)} table section(s):")
            for i, section in enumerate(sections, 1):
                print(f"\n      Section {i}: {section['type'].upper()}")
                print(f"        Start row: {section['start_row']}")
                print(f"        Headers: {section['headers'][:10]}")
                print(f"        Sample data rows:")
                for sample in section['sample_rows'][:3]:
                    print(f"          Row {sample['row']}: {sample['data']}")

        wb.close()

    except Exception as e:
        print(f"  ERROR: {e}")
        import traceback
        traceback.print_exc()

if __name__ == '__main__':
    protocols_dir = Path('/home/user/Growatt_ModbusTCP/Protocols')

    # Analyze VPP file first
    print("\n" + "="*80)
    print("ANALYZING VPP 2.03 PROTOCOL")
    print("="*80)
    vpp_file = protocols_dir / 'GI-BK-E060_GROWATT.VPP.COMMUNICATION.PROTOCOL.OF.INVERTER_V2.03.xlsx'
    if vpp_file.exists():
        analyze_excel_file(vpp_file)
    else:
        print(f"File not found: {vpp_file}")

    # Analyze V1-24 file
    print("\n\n" + "="*80)
    print("ANALYZING V1-24 PROTOCOL")
    print("="*80)
    v124_file = protocols_dir / 'Growatt-Inverter-Modbus-RTU-Protocol-II-V1-24-English-new.xlsx'
    if v124_file.exists():
        analyze_excel_file(v124_file)
    else:
        print(f"File not found: {v124_file}")
